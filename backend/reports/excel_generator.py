"""
Excel Generator — openpyxl.
Genereaza raport Excel profesional cu 4 sheet-uri din verified_data + report_sections.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


# Palette
ACCENT = "6366F1"
WHITE = "FFFFFF"
DARK_BG = "1A1A2E"
LIGHT_GRAY = "F8F9FA"
GREEN = "22C55E"
YELLOW = "EAB308"
RED = "EF4444"

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=ACCENT)
NORMAL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def _style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data_cell(cell, number_format=None):
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")
    if number_format:
        cell.number_format = number_format


def _risk_fill(score):
    if score >= 70:
        return PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    elif score >= 40:
        return PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    else:
        return PatternFill(start_color=RED, end_color=RED, fill_type="solid")


def generate_excel(report_sections: dict, meta: dict, verified_data: dict, output_path: str):
    """Genereaza Excel din verified_data + meta."""
    wb = Workbook()

    # --- Sheet 1: Rezumat ---
    ws1 = wb.active
    ws1.title = "Rezumat"
    ws1.sheet_properties.tabColor = ACCENT

    ws1.merge_cells("A1:D1")
    ws1.cell(row=1, column=1, value=meta.get("title", "Raport RIS")).font = TITLE_FONT
    ws1.merge_cells("A2:D2")
    ws1.cell(row=2, column=1, value=meta.get("company_name", "N/A")).font = Font(name="Calibri", size=12, color="444444")
    ws1.cell(row=4, column=1, value="Generat:").font = Font(bold=True)
    ws1.cell(row=4, column=2, value=meta.get("generated_at", ""))
    ws1.cell(row=5, column=1, value="Nivel raport:").font = Font(bold=True)
    ws1.cell(row=5, column=2, value=meta.get("report_level", "N/A"))
    ws1.cell(row=6, column=1, value="Surse utilizate:").font = Font(bold=True)
    ws1.cell(row=6, column=2, value=meta.get("sources_count", 0))

    # Scor risc
    risk_score = meta.get("risk_score", "N/A")
    numeric_score = meta.get("numeric_score")
    ws1.cell(row=8, column=1, value="SCOR RISC:").font = Font(name="Calibri", bold=True, size=14)
    risk_cell = ws1.cell(row=8, column=2)
    if numeric_score is not None:
        risk_cell.value = f"{risk_score} ({numeric_score}/100)"
    else:
        risk_cell.value = risk_score
    risk_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    if numeric_score is not None:
        risk_cell.fill = _risk_fill(numeric_score)
    ws1.cell(row=9, column=1, value=meta.get("risk_recommendation", "")).font = Font(italic=True, color="888888")

    # Dimensiuni risc
    risk_data = verified_data.get("risk_score", {})
    dimensions = risk_data.get("dimensions", {})
    if dimensions:
        ws1.cell(row=11, column=1, value="Dimensiune").font = HEADER_FONT
        ws1.cell(row=11, column=1).fill = HEADER_FILL
        ws1.cell(row=11, column=2, value="Scor").font = HEADER_FONT
        ws1.cell(row=11, column=2).fill = HEADER_FILL
        ws1.cell(row=11, column=3, value="Pondere").font = HEADER_FONT
        ws1.cell(row=11, column=3).fill = HEADER_FILL
        for i, (dim_name, dim_data) in enumerate(dimensions.items(), start=12):
            ws1.cell(row=i, column=1, value=dim_name.capitalize())
            score_cell = ws1.cell(row=i, column=2, value=dim_data.get("score", 0))
            score_cell.fill = _risk_fill(dim_data.get("score", 0))
            score_cell.font = Font(bold=True, color=WHITE)
            ws1.cell(row=i, column=3, value=f"{dim_data.get('weight', 0)}%")

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 20

    # --- Sheet 2: Date Financiare ---
    ws2 = wb.create_sheet("Financiar")
    ws2.sheet_properties.tabColor = "2563EB"

    financial_official = verified_data.get("financial", {})
    # Extrage date ANAF Bilant multi-an din official_data (transmis prin verified_data)
    bilant_trend = financial_official.get("trend_financiar", {})
    trend_val = bilant_trend.get("value") if isinstance(bilant_trend, dict) else None

    if isinstance(trend_val, dict) and trend_val:
        # Avem trend multi-an
        ws2.cell(row=1, column=1, value="Date Financiare Oficiale (ANAF Bilant)").font = TITLE_FONT
        ws2.merge_cells("A1:F1")

        # Colectam toti anii
        all_years = set()
        metrics = {}
        for metric_key, metric_data in trend_val.items():
            if isinstance(metric_data, dict) and "values" in metric_data:
                metrics[metric_key] = metric_data
                for v in metric_data["values"]:
                    all_years.add(v["year"])

        years = sorted(all_years)
        if years:
            # Header
            ws2.cell(row=3, column=1, value="Indicator")
            for j, year in enumerate(years, start=2):
                ws2.cell(row=3, column=j, value=year)
            ws2.cell(row=3, column=len(years) + 2, value="Trend")
            _style_header_row(ws2, 3, len(years) + 2)

            row = 4
            for metric_key, metric_data in metrics.items():
                ws2.cell(row=row, column=1, value=metric_data.get("name", metric_key))
                _style_data_cell(ws2.cell(row=row, column=1))
                year_map = {v["year"]: v["value"] for v in metric_data["values"]}
                for j, year in enumerate(years, start=2):
                    val = year_map.get(year)
                    cell = ws2.cell(row=row, column=j, value=val)
                    _style_data_cell(cell, "#,##0")
                growth = metric_data.get("growth_percent")
                trend_cell = ws2.cell(row=row, column=len(years) + 2)
                if growth is not None:
                    trend_cell.value = f"{'+' if growth > 0 else ''}{growth}%"
                    trend_cell.font = Font(bold=True, color=GREEN if growth > 0 else RED)
                else:
                    trend_cell.value = "N/A"
                row += 1

            # Grafic CA
            if len(years) >= 2 and "cifra_afaceri_neta" in metrics:
                chart = BarChart()
                chart.type = "col"
                chart.title = "Evolutie Cifra de Afaceri"
                chart.y_axis.title = "RON"
                chart.x_axis.title = "An"
                chart.style = 10

                ca_row = 4
                for idx, mk in enumerate(metrics.keys()):
                    if mk == "cifra_afaceri_neta":
                        ca_row = 4 + idx
                        break

                data_ref = Reference(ws2, min_col=2, max_col=len(years) + 1, min_row=ca_row, max_row=ca_row)
                cats_ref = Reference(ws2, min_col=2, max_col=len(years) + 1, min_row=3, max_row=3)
                chart.add_data(data_ref, from_rows=True)
                chart.set_categories(cats_ref)
                chart.shape = 4
                ws2.add_chart(chart, f"A{row + 2}")

            ws2.column_dimensions["A"].width = 20
            for j in range(2, len(years) + 3):
                ws2.column_dimensions[get_column_letter(j)].width = 16
    else:
        ws2.cell(row=1, column=1, value="Date financiare indisponibile (ANAF Bilant nu a returnat trend multi-an)").font = Font(italic=True, color="888888")

    # --- Sheet 3: Evaluare Risc ---
    ws3 = wb.create_sheet("Evaluare Risc")
    ws3.sheet_properties.tabColor = RED

    ws3.cell(row=1, column=1, value="Evaluare Risc Detaliata").font = TITLE_FONT
    ws3.merge_cells("A1:C1")

    factors = risk_data.get("factors", [])
    if factors:
        ws3.cell(row=3, column=1, value="Factor de Risc")
        ws3.cell(row=3, column=2, value="Severitate")
        _style_header_row(ws3, 3, 2)

        for i, (factor, severity) in enumerate(factors, start=4):
            ws3.cell(row=i, column=1, value=factor)
            _style_data_cell(ws3.cell(row=i, column=1))
            sev_cell = ws3.cell(row=i, column=2, value=severity)
            sev_fill = {
                "HIGH": PatternFill(start_color=RED, end_color=RED, fill_type="solid"),
                "MEDIUM": PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid"),
                "LOW": PatternFill(start_color="60A5FA", end_color="60A5FA", fill_type="solid"),
                "POSITIVE": PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid"),
            }.get(severity, PatternFill())
            sev_cell.fill = sev_fill
            sev_cell.font = Font(bold=True, color=WHITE)
            sev_cell.alignment = Alignment(horizontal="center")
    else:
        ws3.cell(row=3, column=1, value="Niciun factor de risc identificat").font = Font(color=GREEN)

    # Cross-validation
    cross = verified_data.get("cross_validation", {})
    if cross:
        cv_row = max(4 + len(factors), 6)
        ws3.cell(row=cv_row, column=1, value="Cross-Validare Multi-Sursa").font = Font(bold=True, size=12, color=ACCENT)
        ws3.cell(row=cv_row + 1, column=1, value="Camp")
        ws3.cell(row=cv_row + 1, column=2, value="Surse")
        ws3.cell(row=cv_row + 1, column=3, value="Confidence")
        ws3.cell(row=cv_row + 1, column=4, value="Status")
        _style_header_row(ws3, cv_row + 1, 4)

        for i, (field, cv_data) in enumerate(cross.items(), start=cv_row + 2):
            ws3.cell(row=i, column=1, value=field)
            ws3.cell(row=i, column=2, value=", ".join(cv_data.get("sources", [])))
            ws3.cell(row=i, column=3, value=cv_data.get("confidence", 0))
            ws3.cell(row=i, column=4, value=cv_data.get("status", "N/A"))

    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 16

    # --- Sheet 4: Surse ---
    ws4 = wb.create_sheet("Surse")
    ws4.sheet_properties.tabColor = GREEN

    ws4.cell(row=1, column=1, value="Audit Trail — Surse Utilizate").font = TITLE_FONT
    ws4.merge_cells("A1:D1")

    sources = meta.get("sources", [])
    ws4.cell(row=3, column=1, value="Sursa")
    ws4.cell(row=3, column=2, value="Nivel Trust")
    ws4.cell(row=3, column=3, value="Status")
    _style_header_row(ws4, 3, 3)

    for i, src in enumerate(sources, start=4):
        ws4.cell(row=i, column=1, value=src.get("name", "N/A"))
        _style_data_cell(ws4.cell(row=i, column=1))
        ws4.cell(row=i, column=2, value=f"Nivel {src.get('level', '?')}")
        _style_data_cell(ws4.cell(row=i, column=2))
        status_cell = ws4.cell(row=i, column=3, value=src.get("status", "N/A"))
        _style_data_cell(status_cell)
        if src.get("status") == "OK":
            status_cell.font = Font(color=GREEN, bold=True)

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 10

    # Disclaimer
    disc_row = len(sources) + 6
    ws4.cell(row=disc_row, column=1, value="Disclaimer:").font = Font(bold=True, size=9)
    ws4.cell(row=disc_row + 1, column=1, value="Acest raport a fost generat automat folosind exclusiv date disponibile public din surse verificabile.").font = Font(italic=True, size=8, color="888888")

    # --- Sheet 5: KPI ---
    ws5 = wb.create_sheet("KPI")
    ws5.sheet_properties.tabColor = "8B5CF6"

    ws5.cell(row=1, column=1, value="Indicatori Cheie de Performanta (KPI)").font = TITLE_FONT
    ws5.merge_cells("A1:C1")

    ws5.cell(row=3, column=1, value="KPI")
    ws5.cell(row=3, column=2, value="Valoare")
    ws5.cell(row=3, column=3, value="Observatii")
    _style_header_row(ws5, 3, 3)

    kpi_row = 4
    financial = verified_data.get("financial", {})
    bilant_trend_kpi = financial.get("trend_financiar", {})
    trend_kpi_val = bilant_trend_kpi.get("value") if isinstance(bilant_trend_kpi, dict) else None

    # Extract latest year values from trend data
    ca_latest = None
    profit_latest = None
    capitaluri_latest = None
    angajati_latest = None
    ca_first = None
    ca_years_count = 0

    if isinstance(trend_kpi_val, dict):
        for metric_key, metric_data in trend_kpi_val.items():
            if not isinstance(metric_data, dict) or "values" not in metric_data:
                continue
            values = sorted(metric_data["values"], key=lambda v: v["year"])
            if not values:
                continue
            latest_val = values[-1]["value"]
            first_val = values[0]["value"]

            if metric_key == "cifra_afaceri_neta":
                ca_latest = latest_val
                ca_first = first_val
                ca_years_count = len(values)
            elif metric_key == "profit_net":
                profit_latest = latest_val
            elif metric_key == "capitaluri_proprii":
                capitaluri_latest = latest_val
            elif metric_key == "numar_mediu_angajati":
                angajati_latest = latest_val

    # KPI 1: CAGR CA (Compound Annual Growth Rate)
    ws5.cell(row=kpi_row, column=1, value="CAGR Cifra Afaceri")
    _style_data_cell(ws5.cell(row=kpi_row, column=1))
    if ca_first and ca_latest and ca_years_count >= 2 and ca_first > 0:
        n_years = ca_years_count - 1
        # B16 fix: Handle negative CA — CAGR undefined, use simple growth rate
        if ca_latest > 0:
            cagr = ((ca_latest / ca_first) ** (1 / n_years) - 1) * 100
        else:
            cagr = ((ca_latest - ca_first) / ca_first) * 100
        cagr_cell = ws5.cell(row=kpi_row, column=2, value=round(cagr, 2))
        _style_data_cell(cagr_cell, "0.00\"%\"")
        ws5.cell(row=kpi_row, column=3, value=f"Calculat pe {ca_years_count} ani ({n_years} perioade)")
    else:
        ws5.cell(row=kpi_row, column=2, value="N/A")
        ws5.cell(row=kpi_row, column=3, value="Necesita date multi-an cu CA > 0 in primul an")
    _style_data_cell(ws5.cell(row=kpi_row, column=3))
    kpi_row += 1

    # KPI 2: Profit Margin
    ws5.cell(row=kpi_row, column=1, value="Marja Profit Net (%)")
    _style_data_cell(ws5.cell(row=kpi_row, column=1))
    if ca_latest and profit_latest is not None and ca_latest != 0:
        margin = (profit_latest / ca_latest) * 100
        margin_cell = ws5.cell(row=kpi_row, column=2, value=round(margin, 2))
        _style_data_cell(margin_cell, "0.00\"%\"")
        ws5.cell(row=kpi_row, column=3, value=f"Profit {profit_latest:,.0f} / CA {ca_latest:,.0f}")
    else:
        ws5.cell(row=kpi_row, column=2, value="N/A")
        ws5.cell(row=kpi_row, column=3, value="Necesita CA si profit net nenule")
    _style_data_cell(ws5.cell(row=kpi_row, column=3))
    kpi_row += 1

    # KPI 3: Angajati per 1M CA
    ws5.cell(row=kpi_row, column=1, value="Angajati per 1M RON CA")
    _style_data_cell(ws5.cell(row=kpi_row, column=1))
    if angajati_latest and ca_latest and ca_latest > 0:
        ang_per_m = angajati_latest / (ca_latest / 1_000_000)
        ang_cell = ws5.cell(row=kpi_row, column=2, value=round(ang_per_m, 2))
        _style_data_cell(ang_cell, "0.00")
        ws5.cell(row=kpi_row, column=3, value=f"{angajati_latest} angajati, CA {ca_latest:,.0f} RON")
    else:
        ws5.cell(row=kpi_row, column=2, value="N/A")
        ws5.cell(row=kpi_row, column=3, value="Necesita angajati si CA disponibile")
    _style_data_cell(ws5.cell(row=kpi_row, column=3))
    kpi_row += 1

    # KPI 4: ROE estimate
    ws5.cell(row=kpi_row, column=1, value="ROE Estimat (%)")
    _style_data_cell(ws5.cell(row=kpi_row, column=1))
    if profit_latest is not None and capitaluri_latest and capitaluri_latest != 0:
        roe = (profit_latest / capitaluri_latest) * 100
        roe_cell = ws5.cell(row=kpi_row, column=2, value=round(roe, 2))
        _style_data_cell(roe_cell, "0.00\"%\"")
        ws5.cell(row=kpi_row, column=3, value=f"Profit {profit_latest:,.0f} / Cap. proprii {capitaluri_latest:,.0f}")
    else:
        ws5.cell(row=kpi_row, column=2, value="N/A")
        ws5.cell(row=kpi_row, column=3, value="Necesita profit net si capitaluri proprii nenule")
    _style_data_cell(ws5.cell(row=kpi_row, column=3))

    # D12 fix: Add scoring dimensions to KPI sheet
    kpi_row += 2
    ws5.cell(row=kpi_row, column=1, value="Scor pe Dimensiuni (0-100)").font = Font(bold=True, size=11, color="6366f1")
    ws5.merge_cells(start_row=kpi_row, start_column=1, end_row=kpi_row, end_column=3)
    kpi_row += 1

    ws5.cell(row=kpi_row, column=1, value="Dimensiune")
    ws5.cell(row=kpi_row, column=2, value="Scor")
    ws5.cell(row=kpi_row, column=3, value="Pondere / Confidence")
    _style_header_row(ws5, kpi_row, 3)
    kpi_row += 1

    risk_score = verified_data.get("risk_score", {})
    dims = risk_score.get("dimensions", {})
    dim_labels = {"financiar": "Financiar", "juridic": "Juridic", "fiscal": "Fiscal",
                  "operational": "Operational", "reputational": "Reputational", "piata": "Piata"}
    for dim_key, dim_data in dims.items():
        ws5.cell(row=kpi_row, column=1, value=dim_labels.get(dim_key, dim_key.capitalize()))
        _style_data_cell(ws5.cell(row=kpi_row, column=1))
        score_val = dim_data.get("score", 0)
        score_cell = ws5.cell(row=kpi_row, column=2, value=round(score_val, 1))
        _style_data_cell(score_cell, "0.0")
        if score_val >= 70:
            score_cell.font = Font(bold=True, color="22c55e")
        elif score_val >= 40:
            score_cell.font = Font(bold=True, color="eab308")
        else:
            score_cell.font = Font(bold=True, color="ef4444")
        conf = dim_data.get("confidence", 0)
        ws5.cell(row=kpi_row, column=3, value=f"{dim_data.get('weight', 0)}% pondere | Conf: {conf:.0%}")
        _style_data_cell(ws5.cell(row=kpi_row, column=3))
        kpi_row += 1

    # Total score
    total = risk_score.get("numeric_score")
    if total is not None:
        ws5.cell(row=kpi_row, column=1, value="SCOR TOTAL")
        ws5.cell(row=kpi_row, column=1).font = Font(bold=True, size=11)
        total_cell = ws5.cell(row=kpi_row, column=2, value=round(total, 1))
        total_cell.font = Font(bold=True, size=11)
        color_label = risk_score.get("score", "N/A")
        ws5.cell(row=kpi_row, column=3, value=f"Clasificare: {color_label}")

    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 16
    ws5.column_dimensions["C"].width = 45

    wb.save(output_path)
